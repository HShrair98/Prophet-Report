import requests
import base64
import re
import os
from datetime import datetime, timedelta, timezone
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sendgrid
from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
import base64 as b64
 
# ── Credentials from GitHub Secrets ───────────────────────
USERNAME         = os.environ['CRM_USERNAME']
PASSWORD         = os.environ['CRM_PASSWORD']
SENDGRID_API_KEY = os.environ['SENDGRID_API_KEY']
TO_EMAIL         = ['hs@jfrecycle.com', 'rag@jfrecycle.com', 'jf@jfrecycle.com']
FROM_EMAIL       = os.environ.get('FROM_EMAIL', 'hs@jfrecycle.com')
MASTER_FILE      = 'master_activity.xlsx'
 
BASE_URL = 'https://prophetOnDemand.com/prophet/prophetwebservices/AvtProphetApi/odata'
EXCEL_MAX = 32000
 
token = base64.b64encode(f'{USERNAME}:{PASSWORD}'.encode()).decode()
HEADERS = {'Authorization': f'Basic {token}', 'Accept': 'application/json;odata=verbose'}
 
cutoff_dt = datetime.now(timezone.utc) - timedelta(days=7)
cutoff    = cutoff_dt.strftime('%Y-%m-%dT00:00:00')
print(f'Pulling notes from last 7 days (since {cutoff[:10]})')
 
def safe_str(val):
    if val is None: return ''
    s = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', str(val))
    return (s[:EXCEL_MAX] + '... [truncated]') if len(s) > EXCEL_MAX else s
 
def should_skip_note(body):
    if not body or not body.strip(): return True
    lower = body.lower()
    if 'assigned to' in lower: return True
    if 'removed user' in lower: return True
    return False
 
def parse_note(raw):
    if not raw: return '', '', '', ''
    text   = raw.replace('\r\n', '\n').replace('\r', '\n')
    lines  = text.split('\n')
    header = lines[0]
    body   = '\n'.join(lines[1:]).strip()
    chunks = [c.strip() for c in re.split(r'-{3,}', header) if c.strip()]
    date_author   = chunks[0] if chunks else ''
    activity_type = chunks[1] if len(chunks) > 1 else ''
    m_author = re.search(r'Modified by:\s*(.+)$', date_author, re.IGNORECASE)
    author   = m_author.group(1).strip() if m_author else ''
    m_date   = re.match(r'^(.+?)\s*-\s*Modified by:', date_author, re.IGNORECASE)
    date_str = m_date.group(1).strip() if m_date else ''
    return author, activity_type, date_str, body
 
def week_start(d):
    return d - timedelta(days=d.weekday())
 
# ── Fetch notes ────────────────────────────────────────────
note_filter = f"CreatedDate ge datetime'{cutoff}'"
all_notes, skip = [], 0
while True:
    url = (f'{BASE_URL}/Notes?$top=100&$skip={skip}'
           f'&$filter={requests.utils.quote(note_filter)}&$orderby=CreatedDate desc')
    resp = requests.get(url, headers=HEADERS)
    if resp.status_code == 401: raise Exception('Login failed.')
    if resp.status_code != 200: raise Exception(f'API error {resp.status_code}: {resp.text[:300]}')
    page = resp.json().get('value') or resp.json().get('d', {}).get('results', [])
    if not page: break
    all_notes.extend(page)
    print(f'  ...{len(all_notes)} notes fetched')
    if len(page) < 100: break
    skip += 100
 
print(f'Total notes: {len(all_notes)}')
 
# ── Process notes ──────────────────────────────────────────
weekly_activity = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
opp_cache = {}
 
def get_opp(entity_id):
    if entity_id in opp_cache: return opp_cache[entity_id]
    try:
        r = requests.get(f"{BASE_URL}/OpportunityViews?$filter=Id eq guid'{entity_id}'", headers=HEADERS)
        if r.status_code == 200:
            items = r.json().get('value') or r.json().get('d', {}).get('results', [])
            if items:
                opp_cache[entity_id] = items[0]
                return items[0]
    except: pass
    opp_cache[entity_id] = None
    return None
 
results = []
for i, n in enumerate(all_notes):
    entity_id = n.get('EntityId', '')
    raw       = n.get('Text') or ''
    author, activity_type, date_str, body = parse_note(raw)
    if not author and not body: continue
    if author and author.lower() == 'system': continue
    if should_skip_note(body): continue
    opp      = get_opp(entity_id)
    company  = safe_str(opp.get('CompanyName', '')) if opp else ''
    opp_name = safe_str(opp.get('RecordDescription', '')) if opp else ''
    stage    = safe_str(opp.get('UserDefined02', '')) if opp else ''
    city     = safe_str(opp.get('CompanyBusinessCity', '')) if opp else ''
    state    = safe_str(opp.get('CompanyBusinessState', '')) if opp else ''
    compete  = safe_str(opp.get('UserDefined34', '')) if opp else ''
    print(f'  [{i+1}/{len(all_notes)}] {author} — {activity_type} — {company}')
    api_date  = (n.get('CreatedDate') or '')[:10]
    backdated = False
    if date_str and api_date:
        try:
            header_date = datetime.strptime(date_str.strip().split()[0], '%m/%d/%Y').strftime('%Y-%m-%d')
            backdated   = header_date != api_date
        except: pass
    if author and activity_type and api_date:
        try:
            note_dt = datetime.fromisoformat(api_date)
            ws      = week_start(note_dt)
            weekly_activity[ws][author][activity_type] += 1
        except: pass
    results.append({
        'Company Name':     company,
        'Opportunity Name': opp_name,
        'Stage':            stage,
        'City':             city,
        'State':            state,
        'Competition':      compete,
        'User':             safe_str(author),
        'Date':             safe_str(date_str),
        'Activity Type':    safe_str(activity_type),
        'Note':             safe_str(body),
        '_backdated':       backdated,
        '_api_date':        api_date,
    })
 
print(f'Processed {len(results)} note rows')
 
# ── Shared styles ──────────────────────────────────────────
hdr_fill      = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
hdr_font      = Font(bold=True, color='FFFFFF', size=11)
total_fill    = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
total_font    = Font(bold=True, size=11)
flag_fill     = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
flag_hdr_fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
wrap          = Alignment(vertical='top', wrap_text=True)
mid           = Alignment(horizontal='center', vertical='center')
left_mid      = Alignment(horizontal='left', vertical='center')
thin          = Side(style='thin', color='D1D5DB')
border        = Border(left=thin, right=thin, top=thin, bottom=thin)
alt_fill      = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
 
# ── Build weekly report Excel ──────────────────────────────
wb = Workbook()
 
# Sheet 1 - Notes by Activity
ws1 = wb.active
ws1.title = 'Notes by Activity'
headers         = ['Company Name','Opportunity Name','Stage','City','State','Competition','User','Date','Activity Type','Note']
display_headers = headers + ['Backdated?']
for ci, h in enumerate(display_headers, 1):
    c = ws1.cell(row=1, column=ci, value=h)
    c.font=hdr_font; c.alignment=mid; c.border=border
    c.fill = flag_hdr_fill if h == 'Backdated?' else hdr_fill
for ri, rec in enumerate(results, 2):
    is_backdated = rec.get('_backdated', False)
    row_fill     = flag_fill if is_backdated else (alt_fill if ri % 2 == 0 else None)
    for ci, key in enumerate(headers, 1):
        c = ws1.cell(row=ri, column=ci, value=rec[key])
        c.border=border; c.alignment=wrap
        if row_fill: c.fill=row_fill
    c = ws1.cell(row=ri, column=len(display_headers),
                 value=f'YES — entered {rec["_api_date"]}' if is_backdated else '')
    c.border=border; c.alignment=mid
    if is_backdated:
        c.fill=flag_fill; c.font=Font(bold=True, color='856404', size=11)
    elif row_fill: c.fill=row_fill
cw = {'Company Name':28,'Opportunity Name':32,'Stage':26,'City':18,'State':10,'Competition':22,
      'User':22,'Date':22,'Activity Type':28,'Note':70,'Backdated?':28}
for ci, h in enumerate(display_headers, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = cw.get(h, 20)
ws1.row_dimensions[1].height = 22
for ri in range(2, len(results)+2): ws1.row_dimensions[ri].height = 80
ws1.freeze_panes='A2'; ws1.auto_filter.ref=ws1.dimensions
 
# Sheet 2 - Weekly Activity Summary
ws2 = wb.create_sheet(title='Activity Summary (Weekly)')
all_types = sorted(set(at for wd in weekly_activity.values() for ud in wd.values() for at in ud))
all_weeks = sorted(weekly_activity.keys())
sh = ['Week', 'User'] + all_types + ['TOTAL']
for ci, h in enumerate(sh, 1):
    c = ws2.cell(row=1, column=ci, value=h)
    c.font=hdr_font; c.fill=hdr_fill; c.alignment=mid; c.border=border
ri = 2
for ws_dt in all_weeks:
    week_end   = ws_dt + timedelta(days=6)
    week_label = f"{ws_dt.strftime('%m/%d/%Y')} - {week_end.strftime('%m/%d/%Y')}"
    week_users = sorted(weekly_activity[ws_dt].keys())
    for user in week_users:
        is_alt    = ri % 2 == 0
        row_total = 0
        c = ws2.cell(row=ri, column=1, value=week_label)
        c.font=Font(size=11); c.alignment=left_mid; c.border=border
        if is_alt: c.fill=alt_fill
        c = ws2.cell(row=ri, column=2, value=user)
        c.font=Font(bold=True, size=11); c.alignment=left_mid; c.border=border
        if is_alt: c.fill=alt_fill
        for ci, at in enumerate(all_types, 3):
            count = weekly_activity[ws_dt][user].get(at, 0)
            row_total += count
            c = ws2.cell(row=ri, column=ci, value=count if count > 0 else '')
            c.alignment=mid; c.border=border
            if is_alt: c.fill=alt_fill
        c = ws2.cell(row=ri, column=len(sh), value=row_total)
        c.font=total_font; c.alignment=mid; c.border=border; c.fill=total_fill
        ri += 1
    c = ws2.cell(row=ri, column=1, value=f'WEEK TOTAL — {week_label}')
    c.font=total_font; c.fill=total_fill; c.alignment=left_mid; c.border=border
    ws2.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=2)
    week_grand = 0
    for ci, at in enumerate(all_types, 3):
        wt = sum(weekly_activity[ws_dt][u].get(at, 0) for u in week_users)
        week_grand += wt
        c = ws2.cell(row=ri, column=ci, value=wt if wt > 0 else '')
        c.font=total_font; c.fill=total_fill; c.alignment=mid; c.border=border
    c = ws2.cell(row=ri, column=len(sh), value=week_grand)
    c.font=total_font; c.fill=total_fill; c.alignment=mid; c.border=border
    ri += 1
ws2.column_dimensions['A'].width = 32
ws2.column_dimensions['B'].width = 26
for ci in range(3, len(sh)+1):
    ws2.column_dimensions[get_column_letter(ci)].width = 26
ws2.row_dimensions[1].height = 22
for r in range(2, ri): ws2.row_dimensions[r].height = 22
ws2.freeze_panes='C2'
 
# Duplicate touch detection
from datetime import datetime as dt
touch_events = []
for rec in results:
    api_date = rec.get('_api_date', '')
    raw_date = rec.get('Date', '')
    user     = rec.get('User', '')
    company  = rec.get('Company Name', '')
    atype    = rec.get('Activity Type', '')
    if not user or not company or not api_date: continue
    parsed_dt = None
    for fmt in ['%m/%d/%Y %I:%M %p', '%m/%d/%Y %I:%M:%S %p', '%m/%d/%Y']:
        try:
            parsed_dt = dt.strptime(raw_date.strip(), fmt)
            break
        except: pass
    if parsed_dt is None:
        try: parsed_dt = dt.fromisoformat(api_date)
        except: continue
    touch_events.append((parsed_dt, user, company, atype))
 
touch_events.sort(key=lambda x: (x[1], x[2], x[0]))
duplicates = []
i = 0
while i < len(touch_events):
    first_dt, user, company, atype = touch_events[i]
    group = [touch_events[i]]
    j = i + 1
    while j < len(touch_events):
        next_dt, next_user, next_company = touch_events[j][0], touch_events[j][1], touch_events[j][2]
        if next_user != user or next_company != company: break
        if (next_dt - first_dt).total_seconds() / 60 <= 30:
            group.append(touch_events[j]); j += 1
        else: break
    if len(group) > 1:
        for dup in group[1:]:
            duplicates.append((first_dt, dup[0], user, company, dup[3]))
    i = j if j > i + 1 else i + 1
 
gap = 3
dup_start_row = ri + gap
c = ws2.cell(row=dup_start_row, column=1, value='DUPLICATE TOUCHES (same user, same company, within 30 mins)')
c.font=Font(bold=True, color='FFFFFF', size=12)
c.fill=PatternFill(start_color='B45309', end_color='B45309', fill_type='solid')
c.alignment=left_mid; c.border=border
ws2.merge_cells(start_row=dup_start_row, start_column=1, end_row=dup_start_row, end_column=6)
dup_headers = ['User', 'Company', 'First Touch', 'Duplicate Touch', 'Time Apart (mins)', 'Activity Type']
for ci, h in enumerate(dup_headers, 1):
    c = ws2.cell(row=dup_start_row+1, column=ci, value=h)
    c.font=hdr_font; c.fill=hdr_fill; c.alignment=mid; c.border=border
dup_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
dup_alt  = PatternFill(start_color='FDE68A', end_color='FDE68A', fill_type='solid')
if not duplicates:
    c = ws2.cell(row=dup_start_row+2, column=1, value='No duplicate touches found in this period.')
    c.font=Font(italic=True, size=11); c.alignment=left_mid
    ws2.merge_cells(start_row=dup_start_row+2, start_column=1, end_row=dup_start_row+2, end_column=6)
else:
    for di, (first_dt, dup_dt, user, company, atype) in enumerate(duplicates, dup_start_row+2):
        fill = dup_fill if di % 2 == 0 else dup_alt
        mins_apart = round((dup_dt - first_dt).total_seconds() / 60, 1)
        row_vals = [user, company, first_dt.strftime('%m/%d/%Y %I:%M %p'),
                    dup_dt.strftime('%m/%d/%Y %I:%M %p'), mins_apart, atype]
        for ci, val in enumerate(row_vals, 1):
            c = ws2.cell(row=di, column=ci, value=val)
            c.border=border; c.fill=fill
            c.alignment=mid if ci > 2 else left_mid
        ws2.row_dimensions[di].height = 20
ws2.column_dimensions['A'].width = max(32, ws2.column_dimensions['A'].width)
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 22
ws2.column_dimensions['D'].width = 22
ws2.column_dimensions['E'].width = 20
ws2.column_dimensions['F'].width = 28
print(f'Found {len(duplicates)} duplicate touches')
 
weekly_filename = f'avidian_weekly_{datetime.now(timezone.utc).strftime("%Y-%m-%d")}.xlsx'
wb.save(weekly_filename)
print(f'Saved {weekly_filename}')
 
# ── Update Master File ─────────────────────────────────────
today_dt   = datetime.now(timezone.utc)
week_start_dt = week_start(today_dt - timedelta(days=7))
week_end_dt   = week_start_dt + timedelta(days=6)
week_label    = f"{week_start_dt.strftime('%m/%d/%Y')} - {week_end_dt.strftime('%m/%d/%Y')}"
tab_name      = f"Wk {week_start_dt.strftime('%m-%d-%Y')}"
 
# Load existing master or create new one
if os.path.exists(MASTER_FILE):
    master_wb = load_workbook(MASTER_FILE)
    print(f'Loaded existing master file')
else:
    master_wb = Workbook()
    # Remove default sheet
    master_wb.remove(master_wb.active)
    print(f'Created new master file')
 
# ── Master Sheet 1: All Weeks (stacked) ───────────────────
ALL_WEEKS_SHEET = 'All Weeks'
master_headers = ['Week', 'User'] + all_types + ['TOTAL']
 
if ALL_WEEKS_SHEET not in master_wb.sheetnames:
    mws = master_wb.create_sheet(ALL_WEEKS_SHEET, 0)
    # Write headers
    for ci, h in enumerate(master_headers, 1):
        c = mws.cell(row=1, column=ci, value=h)
        c.font=hdr_font; c.fill=hdr_fill; c.alignment=mid; c.border=border
    mws.freeze_panes = 'C2'
    mws.column_dimensions['A'].width = 32
    mws.column_dimensions['B'].width = 26
    for ci in range(3, len(master_headers)+1):
        mws.column_dimensions[get_column_letter(ci)].width = 22
    next_row = 2
else:
    mws = master_wb[ALL_WEEKS_SHEET]
    # Check if this week already exists — if so remove it before re-adding
    rows_to_delete = []
    for row in mws.iter_rows(min_row=2):
        cell_val = row[0].value
        if cell_val and week_label in str(cell_val):
            rows_to_delete.append(row[0].row)
    for r in reversed(rows_to_delete):
        mws.delete_rows(r)
    # Find next empty row
    next_row = mws.max_row + 1
    if next_row == 2 and mws.cell(row=1, column=1).value is None:
        # Headers missing, re-add
        for ci, h in enumerate(master_headers, 1):
            c = mws.cell(row=1, column=ci, value=h)
            c.font=hdr_font; c.fill=hdr_fill; c.alignment=mid; c.border=border
 
# Append this week's data to All Weeks sheet
for ws_dt in all_weeks:
    w_end      = ws_dt + timedelta(days=6)
    w_label    = f"{ws_dt.strftime('%m/%d/%Y')} - {w_end.strftime('%m/%d/%Y')}"
    week_users = sorted(weekly_activity[ws_dt].keys())
    for user in week_users:
        is_alt    = next_row % 2 == 0
        row_total = 0
        c = mws.cell(row=next_row, column=1, value=w_label)
        c.font=Font(size=11); c.alignment=left_mid; c.border=border
        if is_alt: c.fill=alt_fill
        c = mws.cell(row=next_row, column=2, value=user)
        c.font=Font(bold=True, size=11); c.alignment=left_mid; c.border=border
        if is_alt: c.fill=alt_fill
        for ci, at in enumerate(all_types, 3):
            # Ensure column header exists
            if mws.cell(row=1, column=ci).value != at:
                mws.cell(row=1, column=ci, value=at).font = hdr_font
                mws.cell(row=1, column=ci).fill = hdr_fill
                mws.cell(row=1, column=ci).alignment = mid
                mws.cell(row=1, column=ci).border = border
                mws.column_dimensions[get_column_letter(ci)].width = 22
            count = weekly_activity[ws_dt][user].get(at, 0)
            row_total += count
            c = mws.cell(row=next_row, column=ci, value=count if count > 0 else '')
            c.alignment=mid; c.border=border
            if is_alt: c.fill=alt_fill
        # Total column
        total_ci = len(master_headers)
        c = mws.cell(row=next_row, column=total_ci, value=row_total)
        c.font=total_font; c.alignment=mid; c.border=border; c.fill=total_fill
        mws.row_dimensions[next_row].height = 22
        next_row += 1
    # Week subtotal row
    c = mws.cell(row=next_row, column=1, value=f'WEEK TOTAL — {w_label}')
    c.font=total_font; c.fill=total_fill; c.alignment=left_mid; c.border=border
    try: mws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=2)
    except: pass
    wg = 0
    for ci, at in enumerate(all_types, 3):
        wt = sum(weekly_activity[ws_dt][u].get(at, 0) for u in week_users)
        wg += wt
        c = mws.cell(row=next_row, column=ci, value=wt if wt > 0 else '')
        c.font=total_font; c.fill=total_fill; c.alignment=mid; c.border=border
    c = mws.cell(row=next_row, column=len(master_headers), value=wg)
    c.font=total_font; c.fill=total_fill; c.alignment=mid; c.border=border
    mws.row_dimensions[next_row].height = 22
    next_row += 1
 
mws.auto_filter.ref = f"A1:{get_column_letter(len(master_headers))}1"
 
# ── Master Sheet 2: Tab per week ──────────────────────────
if tab_name in master_wb.sheetnames:
    del master_wb[tab_name]
 
week_ws = master_wb.create_sheet(title=tab_name)
for ci, h in enumerate(master_headers, 1):
    c = week_ws.cell(row=1, column=ci, value=h)
    c.font=hdr_font; c.fill=hdr_fill; c.alignment=mid; c.border=border
 
wri = 2
for ws_dt in all_weeks:
    w_end      = ws_dt + timedelta(days=6)
    w_label    = f"{ws_dt.strftime('%m/%d/%Y')} - {w_end.strftime('%m/%d/%Y')}"
    week_users = sorted(weekly_activity[ws_dt].keys())
    for user in week_users:
        is_alt    = wri % 2 == 0
        row_total = 0
        c = week_ws.cell(row=wri, column=1, value=w_label)
        c.font=Font(size=11); c.alignment=left_mid; c.border=border
        if is_alt: c.fill=alt_fill
        c = week_ws.cell(row=wri, column=2, value=user)
        c.font=Font(bold=True, size=11); c.alignment=left_mid; c.border=border
        if is_alt: c.fill=alt_fill
        for ci, at in enumerate(all_types, 3):
            count = weekly_activity[ws_dt][user].get(at, 0)
            row_total += count
            c = week_ws.cell(row=wri, column=ci, value=count if count > 0 else '')
            c.alignment=mid; c.border=border
            if is_alt: c.fill=alt_fill
        c = week_ws.cell(row=wri, column=len(master_headers), value=row_total)
        c.font=total_font; c.alignment=mid; c.border=border; c.fill=total_fill
        week_ws.row_dimensions[wri].height = 22
        wri += 1
    c = week_ws.cell(row=wri, column=1, value=f'WEEK TOTAL — {w_label}')
    c.font=total_font; c.fill=total_fill; c.alignment=left_mid; c.border=border
    try: week_ws.merge_cells(start_row=wri, start_column=1, end_row=wri, end_column=2)
    except: pass
    wg = 0
    for ci, at in enumerate(all_types, 3):
        wt = sum(weekly_activity[ws_dt][u].get(at, 0) for u in week_users)
        wg += wt
        c = week_ws.cell(row=wri, column=ci, value=wt if wt > 0 else '')
        c.font=total_font; c.fill=total_fill; c.alignment=mid; c.border=border
    c = week_ws.cell(row=wri, column=len(master_headers), value=wg)
    c.font=total_font; c.fill=total_fill; c.alignment=mid; c.border=border
    wri += 1
week_ws.column_dimensions['A'].width = 32
week_ws.column_dimensions['B'].width = 26
for ci in range(3, len(master_headers)+1):
    week_ws.column_dimensions[get_column_letter(ci)].width = 22
week_ws.row_dimensions[1].height = 22
week_ws.freeze_panes = 'C2'
 
master_wb.save(MASTER_FILE)
print(f'Master file updated: {MASTER_FILE}')
 
# ── Commit master file back to GitHub repo ─────────────────
import subprocess
subprocess.run(['git', 'config', 'user.email', 'actions@github.com'], check=True)
subprocess.run(['git', 'config', 'user.name', 'GitHub Actions'], check=True)
subprocess.run(['git', 'add', MASTER_FILE], check=True)
result = subprocess.run(['git', 'diff', '--cached', '--quiet'])
if result.returncode != 0:
    subprocess.run(['git', 'commit', '-m', f'Auto-update master activity file — {week_label}'], check=True)
    subprocess.run(['git', 'push'], check=True)
    print('Master file committed and pushed to GitHub')
else:
    print('No changes to master file')
 
# ── Send email via SendGrid ────────────────────────────────
with open(weekly_filename, 'rb') as f:
    weekly_data = f.read()
with open(MASTER_FILE, 'rb') as f:
    master_data = f.read()
 
today_str    = datetime.now(timezone.utc).strftime('%B %d, %Y')
week_ago_str = (datetime.now(timezone.utc) - timedelta(days=7)).strftime('%B %d, %Y')
 
message = Mail(
    from_email=FROM_EMAIL,
    to_emails=TO_EMAIL,
    subject=f'Avidian CRM Weekly Report — {week_ago_str} to {today_str}',
    html_content=f'''
    <h2>Avidian CRM Weekly Report</h2>
    <p>Your automated weekly CRM report is attached.</p>
    <p><strong>Period:</strong> {week_ago_str} to {today_str}</p>
    <p><strong>Total activities logged:</strong> {len(results)}</p>
    <p style="color:#856404;"><strong>Note:</strong> Rows highlighted in yellow indicate backdated entries.</p>
    <p>The updated master activity file is also attached — it contains all weeks to date.</p>
    <p style="color:#6B7280; font-size:12px;"><strong>Heads up:</strong> Emails sent through Prophet are not included in this report. Avidian's API does not expose the Emails tab, so those activities cannot be pulled automatically.</p>
    <br>
    <p style="color:#888; font-size:12px;">This report was generated automatically every Monday at 7:30am EST.</p>
    '''
)
 
for fname, fdata in [(weekly_filename, weekly_data), (MASTER_FILE, master_data)]:
    attachment = Attachment(
        FileContent(b64.b64encode(fdata).decode()),
        FileName(fname),
        FileType('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
        Disposition('attachment')
    )
    message.add_attachment(attachment)
 
sg = sendgrid.SendGridAPIClient(api_key=SENDGRID_API_KEY)
response = sg.send(message)
print(f'Email sent! Status: {response.status_code}')
